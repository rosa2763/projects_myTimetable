from tkinter import*
from openpyxl import load_workbook
class TimeTable(Toplevel):
	def __init__(self,parent,dept_name="timeTable_ME.xlsx",semester="s1"):
		self.dept_name=dept_name
		self.semester=semester
		super().__init__(parent)
		self.title("Small Time Table")
		self.geometry("800x400")
		self.parent = parent
		self.iconbitmap('Logo.ico')
		#Basic timetable with 5 days in a week and 6 slots
		self.columnconfigure(0,weight=1)
		self.rowconfigure(0,weight=1)
		self.tt=LabelFrame(self,text="  Class Time Table  ",fg="Blue",font="bold", border=5,borderwidth=5,padx=10,pady=10)
		self.tt.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
		self.tt.columnconfigure(0,weight=1)
		self.tt.rowconfigure(0,weight=1)
		self.class_label=Label(self.tt,text="CLASS",bg="red",font="bold",width=10,fg="white",border=5,borderwidth=2,bd=2,padx=5,pady=5)
		self.class_label.grid(row=0,column=0,padx=2,pady=2,sticky="nsew")
		
		self.row0=["SLOT1","SLOT2","SLOT3","SLOT4","SLOT5","SLOT6"]
		for i in range(6):
			self.tt.columnconfigure(i+1,weight=1)
			self.tt.rowconfigure(0,weight=1)
			self.lr0=Label(self.tt,text=self.row0[i],bg="grey",fg="white",font="bold",width=10,border=5,borderwidth=2,bd=2,padx=5,pady=5)
			self.lr0.grid(row=0,column=i+1,padx=2,pady=2,sticky="nsew")

		self.column0=["Monday","Tuesday","Wednesday","Thursday","Friday"]
		for i in range(5):
			self.tt.columnconfigure(0,weight=1)
			self.tt.rowconfigure(i+1,weight=1)
			self.lc0 = Label(self.tt,text=self.column0[i],width=10,bg="grey",fg="white",font="bold", border=5, borderwidth=2, bd=2, padx=5, pady=5)
			self.lc0.grid(row=i+1,column=0,padx=2,pady=2,sticky="nsew")
			
		# Load data from Excel Timetable file
		self.workbook = load_workbook(filename=self.dept_name,read_only=True)
		self.sheet = self.workbook[self.semester]
		self.data = [[self.sheet.cell(row, col).value for col in range(2, self.sheet.max_column+1)] for row in range(2, self.sheet.max_row + 1)]
		self.workbook.close()
		for ro in range(5):
			for co in range(6):
				self.tt.rowconfigure(ro+1,weight=1)
				self.tt.columnconfigure(co+1,weight=1)
				self.lmn=Label(self.tt,text=f"{self.data[ro][co]}",bg="lightblue",width=10,border=5,borderwidth=2,bd=2,padx=5,pady=5)
				self.lmn.grid(row=ro+1,column=co+1,padx=2,pady=2,sticky="nsew")
		#self.close_button = Button(self,text="Close Time Table Window",bg="Yellow",font="bold",command=self.destroy)
		self.close_button = Button(self,text="Close Time Table Window",bg="Yellow",font="bold",command=self.destroy)
		self.close_button.grid(row=1,column=0,padx=10,pady=5,sticky="nsew")

class MainWindow(Tk):
	def __init__(self):
		super().__init__()
		self.title("Main window")
		self.geometry("500x400")
		self.iconbitmap('Logo.ico')
		self.open_tt_button = Button(self,text="Open",command=self.open_tt)
		self.open_tt_button.pack(pady=20)

	def open_tt(self):
		TimeTable(self,'timeTable_CE.xlsx','s1')

if __name__ =="__main__":
	app = MainWindow()
	app.mainloop()